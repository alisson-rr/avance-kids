#!/usr/bin/env python3
"""
Importa os Programas ABA oficiais (planilha da cliente) para uma migration SQL.

Fonte:
    AvanceKids-DOCUMENTACAO/LOGICA-ATUALIZADA/Programas_ABA_Completo_F01A_a_F06A_17_colunas.xlsx
    aba "Consolidado", 17 colunas, 450 linhas = 150 códigos x 3 níveis.

Uso:
    python scripts/import_programas.py --check                 # só valida
    python scripts/import_programas.py --out supabase/migrations/<checklist>.sql
    python scripts/import_programas.py --out-triagem supabase/migrations/<triagem>.sql
    python scripts/import_programas.py --self-check            # testes do próprio script

O script falha (exit != 0) em qualquer inconsistência: código fora do padrão,
nível desconhecido, campo obrigatório vazio, duplicidade de (código, nível),
habilidade divergente entre níveis do mesmo código, ou contagem diferente da
esperada. Nada é gravado quando a validação falha.

Decisões de mapeamento (todas explícitas aqui, nenhuma no SQL gerado):

  * A letra da categoria no código define a habilidade:
        AC -> comunicacao   AS -> social      AG -> cognitiva
        AM -> motora        AF -> funcional
    Confere com o agrupamento do checklist no documento oficial
    ("2026.08.18_Logica App para exercícios.docx").

  * AT (F01AT001..F06AT004) é importado, mas para `screening_programs`, não
    para `exercises`. São os "Programas Básicos de Engajamento" da Triagem
    Inicial e têm as mesmas 17 colunas preenchidas dos demais — é conteúdo
    completo, então guardar é obrigação, não decisão.

    O que NÃO cabe em `exercises`: a tabela exige `skill_id NOT NULL
    REFERENCES skills(id)` e só existem 5 habilidades (comunicacao, social,
    cognitiva, motora, funcional), espelhadas em HabilidadeKey no app. Colocar
    AT ali obrigaria a inventar a habilidade de cada programa — e a coluna
    Função não resolve, porque "Atenção conjunta" aparece tanto em AC
    (comunicação) quanto em AG (cognitiva) no próprio arquivo. Além disso,
    qualquer linha em `exercises` fica elegível para generate-activity-plan.

    `screening_programs` guarda o conteúdo sem skill_id e sem vínculo com
    plano: os dados existem, a regra de quando usá-los (gatilho da triagem,
    rebaixamento) continua pendente da cliente.

  * `ordem` = número sequencial do código dentro da faixa+habilidade
    (F01AC001 -> 1 ... F01AC005 -> 5), o MESMO valor nos três níveis.
    Isso preserva a travessia que check_exercise_completion já faz hoje
    (aquisição -> generalização -> manutenção do mesmo código, depois o
    próximo código). O algoritmo não é alterado.

  * `plano` = 'free' em todas as linhas, igual ao seed atual. Quais atividades
    são premium é decisão comercial pendente; marcar por conta própria mudaria
    o que o assinante recebe.
"""

import argparse
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

FONTE_PADRAO = (
    "AvanceKids-DOCUMENTACAO/LOGICA-ATUALIZADA/"
    "Programas_ABA_Completo_F01A_a_F06A_17_colunas.xlsx"
)
ABA = "Consolidado"

CODIGO_RE = re.compile(r"^F(0[1-6])A([CSGMFT])(\d{3})$")

CATEGORIA_SKILL = {
    "C": "comunicacao",
    "S": "social",
    "G": "cognitiva",
    "M": "motora",
    "F": "funcional",
}
CATEGORIA_TRIAGEM = "T"

NIVEL_ENUM = {
    "aquisicao": "aquisicao",
    "generalizacao": "generalizacao",
    "manutencao": "manutencao",
}

# xlsx -> coluna de `exercises`. `titulo` e `codigo` saem de Código/Habilidade.
COLUNAS = [
    ("Código", "codigo"),
    ("Habilidade", "titulo"),
    ("Programa ABA", "programa_aba"),
    ("Função", "funcao"),
    ("Nível", "_nivel"),
    ("Objetivo", "objetivo"),
    ("Materiais", "materiais"),
    ("Recursos", "recursos_extras"),
    ("Exemplos de Brincadeiras", "brincadeiras"),
    ("Frequência", "frequencia"),
    ("Hierarquia de Dicas", "hierarquia_dicas"),
    ("Procedimento", "procedimento"),
    ("Resposta Esperada", "resposta_esperada"),
    ("Procedimento de Correção", "procedimento_correcao"),
    ("Critério de Avanço", "criterio_avanco"),
    ("Registro de Dados (detalhado)", "registro_dados"),
    ("Exemplos de Reforços", "reforcos"),
]

# Todas as 17 colunas são obrigatórias: a planilha oficial não tem célula vazia
# e uma lacuna significaria arquivo trocado ou corrompido.
OBRIGATORIAS = [xlsx for xlsx, _ in COLUNAS]

ESPERADO = {
    "linhas": 450,
    "codigos": 150,
    "niveis_por_codigo": 3,
    "faixas": 6,
}


class ErroDeValidacao(Exception):
    pass


def normalizar_nivel(valor: str) -> str:
    """'Aquisição' -> 'aquisicao'. Sem acento e sem caixa, como o enum do banco."""
    sem_acento = unicodedata.normalize("NFKD", valor).encode("ascii", "ignore").decode()
    return sem_acento.strip().lower()


def ler_planilha(caminho: Path) -> list[dict]:
    import openpyxl  # importado aqui para o --self-check rodar sem a dependência

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        raise ErroDeValidacao(f"aba '{ABA}' não encontrada em {caminho} (abas: {wb.sheetnames})")

    ws = wb[ABA]
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()

    if not linhas:
        raise ErroDeValidacao(f"{caminho} está vazio")

    cabecalho = [str(c).strip() if c is not None else "" for c in linhas[0]]
    esperado = [xlsx for xlsx, _ in COLUNAS]
    if cabecalho != esperado:
        raise ErroDeValidacao(
            "cabeçalho diferente do esperado.\n"
            f"  esperado: {esperado}\n"
            f"  lido:     {cabecalho}"
        )

    return [dict(zip(cabecalho, linha)) for linha in linhas[1:]]


def validar(registros: list[dict]) -> tuple[list[dict], list[dict], list[str]]:
    """Devolve (checklist, triagem_at, problemas). Não levanta: quem chama decide."""
    problemas: list[str] = []
    checklist: list[dict] = []
    triagem: list[dict] = []

    if len(registros) != ESPERADO["linhas"]:
        problemas.append(
            f"esperava {ESPERADO['linhas']} linhas de dados, encontrei {len(registros)}"
        )

    vistos: set[tuple[str, str]] = set()
    habilidade_por_codigo: dict[str, str] = {}

    for i, reg in enumerate(registros, start=2):  # linha 1 é o cabeçalho
        codigo = str(reg.get("Código") or "").strip()
        match = CODIGO_RE.match(codigo)
        if not match:
            problemas.append(f"linha {i}: código fora do padrão FxxA<C|S|G|M|F|T>nnn: {codigo!r}")
            continue

        faixa_num, categoria, seq = match.groups()

        vazios = [c for c in OBRIGATORIAS if not str(reg.get(c) or "").strip()]
        if vazios:
            problemas.append(f"linha {i} ({codigo}): campo obrigatório vazio: {', '.join(vazios)}")
            continue

        nivel = normalizar_nivel(str(reg["Nível"]))
        if nivel not in NIVEL_ENUM:
            problemas.append(f"linha {i} ({codigo}): nível desconhecido {reg['Nível']!r}")
            continue

        chave = (codigo, nivel)
        if chave in vistos:
            problemas.append(f"linha {i}: (código, nível) duplicado: {codigo} / {nivel}")
            continue
        vistos.add(chave)

        titulo = str(reg["Habilidade"]).strip()
        anterior = habilidade_por_codigo.setdefault(codigo, titulo)
        if anterior != titulo:
            problemas.append(
                f"linha {i} ({codigo}): Habilidade divergente entre níveis "
                f"({anterior!r} vs {titulo!r})"
            )
            continue

        item = {
            "codigo": codigo,
            "faixa": f"F{faixa_num}A",
            "categoria": categoria,
            "ordem": int(seq),
            "nivel": NIVEL_ENUM[nivel],
            "titulo": titulo,
        }
        for xlsx, coluna in COLUNAS:
            if coluna.startswith("_") or coluna in ("codigo", "titulo"):
                continue
            item[coluna] = str(reg[xlsx]).strip()

        if categoria == CATEGORIA_TRIAGEM:
            triagem.append(item)
        else:
            item["skill_key"] = CATEGORIA_SKILL[categoria]
            checklist.append(item)

    # Cada código precisa ter exatamente os três níveis.
    niveis_por_codigo = defaultdict(set)
    for item in checklist + triagem:
        niveis_por_codigo[item["codigo"]].add(item["nivel"])

    for codigo, niveis in sorted(niveis_por_codigo.items()):
        if len(niveis) != ESPERADO["niveis_por_codigo"]:
            problemas.append(
                f"{codigo}: esperava {ESPERADO['niveis_por_codigo']} níveis, "
                f"encontrei {sorted(niveis)}"
            )

    if len(niveis_por_codigo) != ESPERADO["codigos"]:
        problemas.append(
            f"esperava {ESPERADO['codigos']} códigos distintos, "
            f"encontrei {len(niveis_por_codigo)}"
        )

    faixas = {item["faixa"] for item in checklist + triagem}
    if len(faixas) != ESPERADO["faixas"]:
        problemas.append(f"esperava {ESPERADO['faixas']} faixas, encontrei {sorted(faixas)}")

    # (faixa, habilidade, ordem, nível) tem que ser único: é a chave que define
    # a travessia do plano. Colisão aqui embaralharia a trilha da criança.
    colisoes = Counter(
        (i["faixa"], i["skill_key"], i["ordem"], i["nivel"]) for i in checklist
    )
    for chave, n in sorted(colisoes.items()):
        if n > 1:
            problemas.append(f"colisão de (faixa, habilidade, ordem, nível): {chave} aparece {n}x")

    return checklist, triagem, problemas


def resumo(checklist: list[dict], triagem: list[dict]) -> str:
    linhas = []
    por_faixa = Counter(i["faixa"] for i in checklist)
    por_skill = Counter(i["skill_key"] for i in checklist)
    por_nivel = Counter(i["nivel"] for i in checklist)

    linhas.append(f"checklist (AC/AS/AG/AM/AF): {len(checklist)} linhas, "
                  f"{len({i['codigo'] for i in checklist})} códigos")
    linhas.append(f"  por faixa:      {dict(sorted(por_faixa.items()))}")
    linhas.append(f"  por habilidade: {dict(sorted(por_skill.items()))}")
    linhas.append(f"  por nível:      {dict(sorted(por_nivel.items()))}")
    linhas.append(f"triagem AT -> screening_programs: {len(triagem)} linhas, "
                  f"{len({i['codigo'] for i in triagem})} códigos")
    linhas.append(f"TOTAL: {len(checklist) + len(triagem)} linhas, "
                  f"{len({i['codigo'] for i in checklist + triagem})} códigos")
    return "\n".join(linhas)


def sql_literal(valor: str | None) -> str:
    if valor is None:
        return "NULL"
    return "'" + str(valor).replace("'", "''") + "'"


CAMPOS_SQL = [
    "skill_key", "faixa", "codigo", "titulo", "nivel", "ordem",
    "programa_aba", "funcao", "objetivo", "procedimento", "materiais",
    "recursos_extras", "frequencia", "brincadeiras", "hierarquia_dicas",
    "resposta_esperada", "procedimento_correcao", "criterio_avanco",
    "registro_dados", "reforcos",
]


def gerar_sql(checklist: list[dict], triagem: list[dict], fonte: Path) -> str:
    # Ordem determinística: mesma entrada, mesmo arquivo, byte a byte.
    ordem_nivel = {"aquisicao": 0, "generalizacao": 1, "manutencao": 2}
    linhas = sorted(
        checklist,
        key=lambda i: (i["faixa"], i["skill_key"], i["ordem"], ordem_nivel[i["nivel"]]),
    )

    valores = []
    for item in linhas:
        campos = []
        for campo in CAMPOS_SQL:
            valor = item[campo]
            campos.append(str(valor) if campo == "ordem" else sql_literal(valor))
        valores.append("    (" + ", ".join(campos) + ")")

    codigos_at = sorted({i["codigo"] for i in triagem})
    total_linhas = len(linhas)
    total_codigos = len({i["codigo"] for i in linhas})
    primeiro_at = codigos_at[0] if codigos_at else "-"
    ultimo_at = codigos_at[-1] if codigos_at else "-"

    return f"""-- ============================================================
-- migration-08: conteúdo oficial dos Programas ABA
--
-- GERADO POR scripts/import_programas.py — NÃO EDITAR À MÃO.
-- Fonte: {fonte.as_posix()}
--        (aba "{ABA}", {ESPERADO['linhas']} linhas, {ESPERADO['codigos']} códigos x 3 níveis)
--
-- O que esta migration faz:
--   1. adiciona `programa_aba` e `funcao` em exercises — as duas colunas da
--      planilha oficial que não tinham destino no schema;
--   2. aposenta as atividades genéricas semeadas por migration-03
--      (o próprio arquivo se declara "Conteúdo genérico de MVP");
--   3. insere {total_linhas} atividades oficiais ({total_codigos} códigos x 3 níveis).
--
-- O que esta migration NÃO faz:
--   * não altera generate-activity-plan, check_exercise_completion,
--     resolve_age_bracket nem qualquer regra de recomendação/progressão;
--   * não coloca em `exercises` os {len(codigos_at)} códigos de triagem AT
--     ({primeiro_at}..{ultimo_at}) — são os Programas Básicos de
--     Engajamento e vão para `screening_programs` na migration-10, sem
--     habilidade e sem vínculo com plano: o conteúdo existe, a regra de
--     disparo continua pendente de definição da cliente;
--   * não marca nenhuma atividade como premium (todas entram como 'free',
--     igual ao seed anterior).
--
-- Compatibilidade com dados existentes:
--   As atividades antigas ainda referenciadas por algum activity_plans são
--   arquivadas em vez de removidas, para não violar a FK e não apagar o
--   histórico de nenhuma criança. Crianças que já tinham plano continuam
--   apontando para a atividade antiga (arquivada) e só passam a ver o
--   conteúdo oficial ao refazer a triagem, que regera o plano.
-- ============================================================

-- ── 1. Colunas novas ────────────────────────────────────────
ALTER TABLE exercises ADD COLUMN IF NOT EXISTS programa_aba TEXT;
ALTER TABLE exercises ADD COLUMN IF NOT EXISTS funcao TEXT;

COMMENT ON COLUMN exercises.programa_aba IS 'Nome do programa ABA na planilha oficial (coluna "Programa ABA").';
COMMENT ON COLUMN exercises.funcao IS 'Função comportamental alvo na planilha oficial (coluna "Função").';

-- ── 2. Aposentar o conteúdo placeholder ─────────────────────
-- O seed de migration-03 gerava o código como `<faixa>-<SKI>-<n>` (ex.:
-- 'F01A-COM-01'); o código oficial não tem hífen (ex.: 'F01AC001'). O padrão
-- com hífen isola exatamente as linhas do placeholder.
DELETE FROM exercises
WHERE codigo ~ '^F0[1-6]A-'
  AND NOT EXISTS (SELECT 1 FROM activity_plans ap WHERE ap.exercise_id = exercises.id);

UPDATE exercises SET status = 'arquivado'
WHERE codigo ~ '^F0[1-6]A-' AND status <> 'arquivado';

-- ── 3. Conteúdo oficial ─────────────────────────────────────
WITH oficial(
  skill_key, faixa_codigo, codigo, titulo, nivel, ordem,
  programa_aba, funcao, objetivo, procedimento, materiais,
  recursos_extras, frequencia, brincadeiras, hierarquia_dicas,
  resposta_esperada, procedimento_correcao, criterio_avanco,
  registro_dados, reforcos
) AS (
  VALUES
{",\n".join(valores)}
)
INSERT INTO exercises (
  skill_id, age_bracket_id, codigo, titulo, media_type, nivel, ordem, plano, status,
  programa_aba, funcao, objetivo, procedimento, materiais, recursos_extras,
  frequencia, brincadeiras, hierarquia_dicas, resposta_esperada,
  procedimento_correcao, criterio_avanco, registro_dados, reforcos
)
SELECT
  s.id,
  b.id,
  o.codigo,
  o.titulo,
  'imagem'::media_type,
  o.nivel::exercise_level,
  o.ordem,
  'free'::subscription_plan,
  'ativo'::record_status,
  o.programa_aba, o.funcao, o.objetivo, o.procedimento, o.materiais,
  o.recursos_extras, o.frequencia, o.brincadeiras, o.hierarquia_dicas,
  o.resposta_esperada, o.procedimento_correcao, o.criterio_avanco,
  o.registro_dados, o.reforcos
FROM oficial o
JOIN skills s ON s.key = o.skill_key
JOIN age_brackets b ON b.codigo = o.faixa_codigo;

-- ── 4. Trava de duplicidade ─────────────────────────────────
-- `exercises` não tinha nenhuma restrição de unicidade, então reaplicar um
-- importador duplicaria a trilha inteira da criança em silêncio. Índice
-- parcial porque `codigo` é opcional para atividades criadas pelo backoffice.
CREATE UNIQUE INDEX IF NOT EXISTS uq_exercises_codigo_nivel
  ON exercises (codigo, nivel) WHERE codigo IS NOT NULL;

-- ── 5. Conferência ──────────────────────────────────────────
-- Falha a migration inteira se o JOIN tiver perdido linha (skill.key ou
-- age_brackets.codigo divergente do esperado) — melhor não aplicar do que
-- aplicar pela metade.
DO $$
DECLARE
  v_total INTEGER;
BEGIN
  SELECT count(*) INTO v_total FROM exercises WHERE codigo ~ '^F0[1-6]A[CSGMF]\\d{{3}}$';
  IF v_total <> {total_linhas} THEN
    RAISE EXCEPTION 'esperava {total_linhas} atividades oficiais, encontrei %', v_total;
  END IF;
END $$;
"""


CAMPOS_SQL_TRIAGEM = [
    "faixa", "codigo", "titulo", "nivel", "ordem",
    "programa_aba", "funcao", "objetivo", "procedimento", "materiais",
    "recursos_extras", "frequencia", "brincadeiras", "hierarquia_dicas",
    "resposta_esperada", "procedimento_correcao", "criterio_avanco",
    "registro_dados", "reforcos",
]


def gerar_sql_triagem(triagem: list[dict], fonte: Path) -> str:
    """Migration dos Programas Básicos de Engajamento (códigos AT).

    Tabela própria de propósito: ver a nota sobre AT no topo do arquivo.
    """
    ordem_nivel = {"aquisicao": 0, "generalizacao": 1, "manutencao": 2}
    linhas = sorted(triagem, key=lambda i: (i["faixa"], i["ordem"], ordem_nivel[i["nivel"]]))

    valores = []
    for item in linhas:
        campos = []
        for campo in CAMPOS_SQL_TRIAGEM:
            valor = item[campo]
            campos.append(str(valor) if campo == "ordem" else sql_literal(valor))
        valores.append("    (" + ", ".join(campos) + ")")

    codigos = sorted({i["codigo"] for i in linhas})
    total_linhas = len(linhas)
    total_codigos = len(codigos)

    return f"""-- ============================================================
-- migration-10: Programas Básicos de Engajamento (códigos AT)
--
-- GERADO POR scripts/import_programas.py — NÃO EDITAR À MÃO.
-- Fonte: {fonte.as_posix()}
--        (aba "{ABA}", mesmas 17 colunas dos demais programas)
--
-- POR QUE UMA TABELA SEPARADA E NÃO `exercises`
--
-- Os {total_codigos} códigos AT ({codigos[0]}..{codigos[-1]}) têm as 17 colunas
-- preenchidas, exatamente como os {ESPERADO['codigos'] - total_codigos} códigos do
-- checklist: são programas completos e precisam existir no banco.
--
-- Só que `exercises.skill_id` é NOT NULL REFERENCES skills(id) e o catálogo
-- tem exatamente 5 habilidades (baseline.sql:877 — comunicacao, social,
-- cognitiva, motora, funcional), espelhadas em HabilidadeKey no app. Inserir
-- AT ali exigiria decidir a habilidade de cada programa, e a coluna "Função"
-- não resolve: "Atenção conjunta" aparece tanto em AC quanto em AG na própria
-- planilha. Pior: toda linha de `exercises` é elegível para
-- generate-activity-plan, então o conteúdo entraria no plano da criança sem
-- que a regra de disparo estivesse definida.
--
-- Esta tabela guarda o conteúdo SEM skill_id, SEM `plano` e SEM qualquer
-- vínculo com activity_plans. Consequência: os dados passam a existir e podem
-- ser revisados no backoffice; a decisão de QUANDO usá-los (gatilho da
-- Triagem Inicial, rebaixamento de faixa) continua pendente da cliente e não
-- é antecipada por nenhuma linha deste arquivo.
--
-- Quando a regra for definida, o vínculo com a habilidade e com o plano entra
-- em uma migration nova; nada aqui precisa ser reescrito.
-- ============================================================

CREATE TABLE IF NOT EXISTS screening_programs (
  id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
  codigo TEXT NOT NULL,
  age_bracket_id UUID NOT NULL REFERENCES age_brackets(id),
  titulo TEXT NOT NULL,
  nivel exercise_level NOT NULL,
  ordem INTEGER NOT NULL,

  -- Mesmas colunas de conteúdo de `exercises`, na mesma ordem da planilha.
  programa_aba TEXT,
  funcao TEXT,
  objetivo TEXT,
  procedimento TEXT,
  materiais TEXT,
  recursos_extras TEXT,
  frequencia TEXT,
  brincadeiras TEXT,
  hierarquia_dicas TEXT,
  resposta_esperada TEXT,
  procedimento_correcao TEXT,
  criterio_avanco TEXT,
  registro_dados TEXT,
  reforcos TEXT,

  status record_status NOT NULL DEFAULT 'ativo',
  created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
  updated_at TIMESTAMPTZ NOT NULL DEFAULT now(),

  CONSTRAINT uq_screening_programs_codigo_nivel UNIQUE (codigo, nivel)
);

COMMENT ON TABLE screening_programs IS
  'Programas Básicos de Engajamento (códigos AT) da planilha oficial. Conteúdo armazenado sem habilidade e sem vínculo com plano: a regra de utilização depende de definição da cliente.';
COMMENT ON COLUMN screening_programs.ordem IS
  'Sequência do código dentro da faixa (F01AT001 -> 1). Mesmo valor nos três níveis.';

CREATE INDEX IF NOT EXISTS idx_screening_programs_faixa
  ON screening_programs (age_bracket_id, ordem, nivel);

DROP TRIGGER IF EXISTS trg_screening_programs_updated_at ON screening_programs;
CREATE TRIGGER trg_screening_programs_updated_at
  BEFORE UPDATE ON screening_programs
  FOR EACH ROW EXECUTE FUNCTION set_updated_at();

-- RLS: só admin, por enquanto. O app não tem tela para este conteúdo porque
-- não há regra de quando exibi-lo; abrir a leitura para `authenticated` antes
-- disso entregaria conteúdo sem contexto. A policy de leitura do app entra
-- junto com a regra.
ALTER TABLE screening_programs ENABLE ROW LEVEL SECURITY;

CREATE POLICY "Admins manage screening programs"
  ON screening_programs FOR ALL
  TO authenticated
  USING (is_admin())
  WITH CHECK (is_admin());

-- ── Conteúdo oficial ────────────────────────────────────────
WITH triagem(
  faixa_codigo, codigo, titulo, nivel, ordem,
  programa_aba, funcao, objetivo, procedimento, materiais,
  recursos_extras, frequencia, brincadeiras, hierarquia_dicas,
  resposta_esperada, procedimento_correcao, criterio_avanco,
  registro_dados, reforcos
) AS (
  VALUES
{",\n".join(valores)}
)
INSERT INTO screening_programs (
  codigo, age_bracket_id, titulo, nivel, ordem,
  programa_aba, funcao, objetivo, procedimento, materiais, recursos_extras,
  frequencia, brincadeiras, hierarquia_dicas, resposta_esperada,
  procedimento_correcao, criterio_avanco, registro_dados, reforcos
)
SELECT
  t.codigo,
  b.id,
  t.titulo,
  t.nivel::exercise_level,
  t.ordem,
  t.programa_aba, t.funcao, t.objetivo, t.procedimento, t.materiais,
  t.recursos_extras, t.frequencia, t.brincadeiras, t.hierarquia_dicas,
  t.resposta_esperada, t.procedimento_correcao, t.criterio_avanco,
  t.registro_dados, t.reforcos
FROM triagem t
JOIN age_brackets b ON b.codigo = t.faixa_codigo
ON CONFLICT (codigo, nivel) DO NOTHING;

-- ── Conferência ─────────────────────────────────────────────
DO $$
DECLARE
  v_total INTEGER;
  v_vinculo INTEGER;
BEGIN
  SELECT count(*) INTO v_total FROM screening_programs;
  IF v_total <> {total_linhas} THEN
    RAISE EXCEPTION 'esperava {total_linhas} programas de triagem, encontrei %', v_total;
  END IF;

  -- Nenhum código AT pode ter vazado para `exercises`: é lá que
  -- generate-activity-plan procura conteúdo para a criança.
  SELECT count(*) INTO v_vinculo FROM exercises WHERE codigo ~ '^F0[1-6]AT[0-9]{{3}}$';
  IF v_vinculo <> 0 THEN
    RAISE EXCEPTION 'códigos AT não podem estar em exercises (achei %)', v_vinculo;
  END IF;
END $$;
"""


def self_check() -> None:
    """Checagem executável do script, sem tocar na planilha."""
    assert normalizar_nivel("Aquisição") == "aquisicao"
    assert normalizar_nivel(" Generalização ") == "generalizacao"
    assert normalizar_nivel("Manutenção") == "manutencao"

    assert CODIGO_RE.match("F01AC001").groups() == ("01", "C", "001")
    assert CODIGO_RE.match("F06AT004").groups() == ("06", "T", "004")
    assert CODIGO_RE.match("F07AC001") is None, "faixa fora de 01..06 deve falhar"
    assert CODIGO_RE.match("F01AX001") is None, "categoria desconhecida deve falhar"
    assert CODIGO_RE.match("F01A-COM-01") is None, "código placeholder deve falhar"

    assert sql_literal("d'água") == "'d''água'"
    assert sql_literal(None) == "NULL"

    def linha(codigo, nivel, titulo="Habilidade X", **extra):
        base = {c: f"valor {c}" for c in OBRIGATORIAS}
        base["Código"] = codigo
        base["Nível"] = nivel
        base["Habilidade"] = titulo
        base.update(extra)
        return base

    niveis = ["Aquisição", "Generalização", "Manutenção"]

    # Feliz: 1 código completo (as contagens globais reclamam, o resto não).
    _, _, problemas = validar([linha("F01AC001", n) for n in niveis])
    assert not any("duplicado" in p or "divergente" in p or "vazio" in p for p in problemas), problemas

    # Duplicidade de (código, nível).
    _, _, problemas = validar([linha("F01AC001", "Aquisição")] * 2)
    assert any("duplicado" in p for p in problemas), problemas

    # Habilidade divergente entre níveis do mesmo código.
    _, _, problemas = validar([
        linha("F01AC001", "Aquisição", titulo="A"),
        linha("F01AC001", "Generalização", titulo="B"),
    ])
    assert any("divergente" in p for p in problemas), problemas

    # Campo obrigatório vazio.
    _, _, problemas = validar([linha("F01AC001", "Aquisição", Objetivo="  ")])
    assert any("obrigatório vazio" in p for p in problemas), problemas

    # Código fora do padrão.
    _, _, problemas = validar([linha("XPTO", "Aquisição")])
    assert any("fora do padrão" in p for p in problemas), problemas

    # Nível desconhecido.
    _, _, problemas = validar([linha("F01AC001", "Consolidação")])
    assert any("nível desconhecido" in p for p in problemas), problemas

    # Código incompleto (2 níveis em vez de 3).
    _, _, problemas = validar([linha("F01AC001", n) for n in niveis[:2]])
    assert any("esperava 3 níveis" in p for p in problemas), problemas

    # AT vai para a lista de triagem e não entra no checklist.
    checklist, triagem, _ = validar([linha("F01AT001", n) for n in niveis])
    assert checklist == [] and len(triagem) == 3, (checklist, triagem)

    # O SQL da triagem grava em screening_programs e nao pode carregar
    # habilidade nem tocar no plano da crianca — e o que mantem a decisao
    # pedagogica em aberto.
    sql_triagem = gerar_sql_triagem(triagem, Path("planilha.xlsx"))
    assert "INSERT INTO screening_programs" in sql_triagem
    # Comentario explica por que skill_id nao cabe; o SQL executavel nao
    # pode conter nem a coluna nem o vinculo com o plano.
    executavel = "".join(l for l in sql_triagem.splitlines(True) if not l.lstrip().startswith("--"))
    assert "skill_id" not in executavel, "triagem nao pode carregar habilidade"
    assert "activity_plans" not in executavel, "triagem nao pode tocar o plano"
    assert sql_triagem.count("\n    (") == 3, "esperava 3 linhas de VALUES"

    # Categoria -> habilidade.
    checklist, _, _ = validar([linha("F03AG002", "Aquisição")])
    assert checklist[0]["skill_key"] == "cognitiva"
    assert checklist[0]["faixa"] == "F03A"
    assert checklist[0]["ordem"] == 2

    print("self-check: OK")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--fonte", default=FONTE_PADRAO, help="caminho do .xlsx oficial")
    parser.add_argument("--out", help="arquivo .sql do checklist (exercises)")
    parser.add_argument("--out-triagem", help="arquivo .sql dos programas AT (screening_programs)")
    parser.add_argument("--check", action="store_true", help="apenas valida e resume")
    parser.add_argument("--self-check", action="store_true", help="roda os testes do script")
    args = parser.parse_args()

    if args.self_check:
        self_check()
        return 0

    fonte = Path(args.fonte)
    if not fonte.exists():
        print(f"ERRO: arquivo fonte não encontrado: {fonte}", file=sys.stderr)
        return 2

    try:
        registros = ler_planilha(fonte)
    except ErroDeValidacao as err:
        print(f"ERRO: {err}", file=sys.stderr)
        return 2

    checklist, triagem, problemas = validar(registros)

    print(resumo(checklist, triagem))

    if problemas:
        print(f"\n{len(problemas)} PROBLEMA(S) — nada foi gerado:", file=sys.stderr)
        for p in problemas:
            print(f"  - {p}", file=sys.stderr)
        return 1

    print("\nvalidação: OK")

    if args.check or not (args.out or args.out_triagem):
        return 0

    def escrever(caminho: str, conteudo: str) -> None:
        destino = Path(caminho)
        destino.parent.mkdir(parents=True, exist_ok=True)
        destino.write_text(conteudo, encoding="utf-8", newline="\n")
        print(f"gerado: {destino} ({destino.stat().st_size} bytes)")

    if args.out:
        escrever(args.out, gerar_sql(checklist, triagem, fonte))
    if args.out_triagem:
        escrever(args.out_triagem, gerar_sql_triagem(triagem, fonte))
    return 0


if __name__ == "__main__":
    sys.exit(main())
